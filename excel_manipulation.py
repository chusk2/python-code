import os
import openpyxl as excel
# import module for Font styling
from openpyxl.styles import Font
from openpyxl.styles import Alignment
### Change directory to ~/Documentos/EXCEL
os.chdir('/home/daniel/Documentos/EXCEL')
month = 'Enero'
year = '2021'
### Diccionario con los nombres de los alumnos
alumnos = {'Antonio Martin Herrera':'Susana',
            'Fernando Linares Parrado':'Olga',
            'Francisco Jose Ranea Rodriguez':'Adrián',
            'Juan Manuel Rodriguez Marin':'María Rodríguez',
            'Maria Luz Jimenez Callejon':'María Rodríguez',
            'Maria Jimenez Labrador': 'Irene Quesada',
            'Maria Julia Medina Achirica':'Natalia Peña',
            'Maria Teresa Galache Laza':'Edu',
            'Marta Gaspar Luque':'Lucía',
            'Nuria Juarez Lopez':'Carlos',
            'Paloma Infante Marquez':'Blanca',
            'Sonia Lopez Lopez':'Raúl',
            'Yolanda Gonzalez Camara':'Ying ying',
            'Maria Teresa Molina Vilches':'Juan',
           'Javier Sanchez Garcia':'María Sánchez'
           }
## Create a copy of alumnos dict and clear its values
importes_por_alumno = {}
for k,v in alumnos.items() :
    importes_por_alumno.update({v:[k,0]})

# Get the number of first row with content
unicaja_wb = excel.load_workbook('unicaja_enero_21.xlsx')
unicaja_sheet = unicaja_wb.active
last_row = unicaja_sheet.max_row

## Get the first non-empty row of book
for i in range(1,last_row + 1 ) :
    # If cell is empty, it has value of None
    if unicaja_sheet.cell(row=i, column = 1).value != None :
        first_row = i
        break
print(f'The first non-empty row is {first_row}')

for row in range(first_row+1,last_row+1) :
    # Define importe de la fila
    unicaja_importe = int(unicaja_sheet.cell(row,column = 4).value)
    unicaja_concepto = unicaja_sheet.cell(row,column = 3).value.title()
    if unicaja_importe > 0 and (unicaja_concepto in alumnos) :
        # saca nombre del alumno
        alumno = alumnos[unicaja_concepto]
        # importes_por_alumno. Estructura del diccionario:
        # { nombre_alumno : [nombre_padres, importe] } 
        importes_por_alumno[alumno][1] += unicaja_importe
print('Los importes han sido clasificados por alumno.')

### Create a new workbook with classified incomes
# Create a new book
libro_ingresos = excel.Workbook()
# rename the active sheet
libro_ingresos.active.title='Enero 2021'

# active sheet
hoja_ingresos = libro_ingresos.active
# empezamos en la fila 2
alumno_cell = hoja_ingresos['A2']
importe_cell = hoja_ingresos['B2']
padres_cell = hoja_ingresos['C2']
# Set alignment to horizontal
alumno_cell.alignment = Alignment(horizontal='center')
importe_cell.alignment = Alignment(horizontal='center')
padres_cell.alignment = Alignment(horizontal='center')
# Set the name of the cells
alumno_cell.value = 'Alumno'
padres_cell.value = 'Nombre padres'
importe_cell.value = 'Importe (€)'

# Set the title as merged cells in row 1
hoja_ingresos.merge_cells('A1:C1')
header = hoja_ingresos.cell(1,1)
header.value = f'{month} - {year}'
hoja_ingresos.row_dimensions[1].height = 28
# set the font for title
font_header = Font(size=24,bold=True)
header.font = font_header
header.alignment = Alignment(horizontal='center')
# primer fila con ingresos es la fila 3
row_incomes = 3
# importes_por_alumno. Estructura del diccionario:
# { nombre_alumno : [nombre_padres, importe] } 
for k,v in importes_por_alumno.items() :
    hoja_ingresos[f'A{row_incomes}'] = k
    hoja_ingresos[f'B{row_incomes}'] = v[1]
    hoja_ingresos[f'C{row_incomes}'] = v[0]
    # incrementa una fila en la hoja del libro de ingresos
    row_incomes += 1

# Añade fila con los totales
importes = list ( importes_por_alumno.values() )
total_ingresos = sum ( [ i[1] for i in importes ] )
print(total_ingresos)
hoja_ingresos[f'A{row_incomes}'] = 'Total'
hoja_ingresos[f'B{row_incomes}'] = total_ingresos

### Format the cells of the workbook

max_alumno_len = max ( [len(alumno) for alumno in list(alumnos.values()) ] )
max_padres_len = max ( [len(padres) for padres in list(alumnos.keys()) ] )

print(max_alumno_len,max_padres_len)
hoja_ingresos.column_dimensions['A'].width = max_alumno_len
hoja_ingresos.column_dimensions['B'].width = len('Importe (€)')
hoja_ingresos.column_dimensions['C'].width = max_padres_len


# save the book
nombre_libro_ingresos = 'ingresos_enero_21'
libro_ingresos.save(f'{nombre_libro_ingresos}.xlsx')
print(f'Libro guardado con el nombre: {nombre_libro_ingresos}.xlsx')



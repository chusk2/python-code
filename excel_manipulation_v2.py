import os
import openpyxl as excel
# import module for Font styling
from openpyxl.styles import Font
from openpyxl.styles import Alignment
### Change directory to ~/Documentos/EXCEL
os.chdir('/home/daniel/Documentos/EXCEL')
month = 'Enero'
year = '2021'
### Diccionario con los nombres de los alumnos
ordenantes = {'Antonio Martin Herrera':'Susana',
            'Fernando Linares Parrado':'Olga',
            'Francisco Jose Ranea Rodriguez':'Adrián',
            'Juan Manuel Rodriguez Marin':'María Rodríguez',
            'Maria Luz Jimenez Callejon':'María Rodríguez',
            'Maria Jimenez Labrador': 'Irene Quesada',
            'Maria Julia Medina Achirica':'Natalia Peña',
            'Maria Teresa Galache Laza':'Edu',
            'Marta Gaspar Luque':'Lucía',
            'Nuria Juarez Lopez':'Carlos',
            'Paloma Infante Marquez':'Blanca',
            'Sonia Lopez Lopez':'Raúl',
            'Yolanda Gonzalez Camara':'Ying ying',
            'Maria Teresa Molina Vilches':'Juan',
           'Javier Sanchez Garcia':'María Sánchez'
           }
def get_parents_by_student (alumno_name,lista) :
    """Get the parents name searching within dictionary
        by key"""
    alumnos = list ( lista.values())
    padres = list( lista.keys() )
    position = alumnos.index(alumno_name)
    return padres[position]
### VARIABLES NAMES ###
# ordenantes : diccionario con ordenantes y alumno correspondiente
# importes_alumnos : diccionario con k,v -> alumno:importe
# unicaja_file : nombre archivo con datos de unicaja sin tratar
# last_row : ultima fila con datos
# first_row : primera fila que contiene datos (sin contar encabezados)
# unicaja_importe : importe del apunte de la hoja de unicaja
# unicaja_concepto : concepto del apunte de la hoja unicaja
# libro_ingresos : libro del archivo resultado con ingresos
### END OF VARIABLES NAMES ###

#ordenantes = {'Susana':'Antonio Martin Herrera',
# 'Olga':'Fernando Linares Parrado',
# 'Adrián':'Francisco Jose Ranea Rodriguez',
# 'María Rodríguez':'Juan Manuel Rodriguez Marin',
# 'María Rodríguez':'Maria Luz Jimenez Callejon',
# 'Irene Quesada':'Maria Jimenez Labrador',
# 'Natalia Peña':'Maria Julia Medina Achirica',
# 'Edu':'Maria Teresa Galache Laza',
# 'Lucía':'Marta Gaspar Luque',
# 'Carlos':'Nuria Juarez Lopez',
# 'Blanca':'Paloma Infante Marquez',
# 'Raúl':'Sonia Lopez Lopez',
# 'Ying ying':'Yolanda Gonzalez Camara',
# 'Juan':'Maria Teresa Molina Vilches',
# 'María Sánchez':'Javier Sanchez Garcia'
# }

def extract_data(unicaja_file) :
    
    # create a list with all the student names
    keys = list (ordenantes.values() )
    # set all the alumnos amount to zero
    # use zero to be able to add integers
    importes_dic = dict(zip (keys, [0 for i in range(len(keys))] ) )
    
    # define the workbook and the active sheet
    unicaja_wb = excel.load_workbook(f'{unicaja_file}.xlsx')
    unicaja_sheet = unicaja_wb.active
    last_row = unicaja_sheet.max_row

    ## Get the first non-empty row of book
    for i in range(1,last_row + 1 ) :
        # If cell is empty, it has value of None
        if unicaja_sheet.cell(row=i, column = 1).value != None :
            first_row = i
            break
    # add one row to avoid headers
    first_row += 1

    for row in range(first_row,last_row+1) :
        # Define importe de la fila
        unicaja_importe = int(unicaja_sheet.cell(row,column = 4).value)
        # set to title format the concept of the row
        unicaja_concepto = unicaja_sheet.cell(row,column = 3).value.title()
        # check if row has an income and if it comes from student
        if unicaja_importe > 0 and (unicaja_concepto in ordenantes) :
            # saca nombre del alumno
            alumno = ordenantes[unicaja_concepto]
            # importes_alumnos. Estructura del diccionario:
            # { nombre_alumno : importe } 
            importes_dic[alumno] += unicaja_importe
    
    return importes_dic

##### TO DO ####
# Check if incomes books exists and create it does not
# otherwise, add a new sheet
####

### Create a new workbook with classified incomes
# Create a new book

libro_ingresos = excel.Workbook()
# rename the active sheet
libro_ingresos.active.title='Enero 2021'
def create_sheet(month,year,importes_dic) :
    # create sheet
    title_sheet = f'{month} {year}'
    libro_ingresos.create_sheet(title = title_sheet)
    hoja_ingresos = libro_ingresos[title_sheet]

    ### START FORMATING THE HEADERS ###
    # empezamos en la fila 2
    alumno_cell = hoja_ingresos['A2']
    importe_cell = hoja_ingresos['B2']
    padres_cell = hoja_ingresos['C2']
    # Set alignment to horizontal
    alumno_cell.alignment = Alignment(horizontal='center')
    importe_cell.alignment = Alignment(horizontal='center')
    padres_cell.alignment = Alignment(horizontal='center')
    # Set the name of the cells
    alumno_cell.value = 'Alumno'
    importe_cell.value = 'Importe (€)'
    padres_cell.value = 'Nombre padres'

    # Set the title as merged cells in row 1
    hoja_ingresos.merge_cells('A1:C1')
    header = hoja_ingresos.cell(1,1)
    header.value = f'{month} - {year}'
    hoja_ingresos.row_dimensions[1].height = 28
    # set the font for title
    font_header = Font(size=24,bold=True)
    header.font = font_header
    header.alignment = Alignment(horizontal='center')
    ### END OF FORMAT OF HEADERS

    # data to populate sheet start at row 3
    row = 3
    # importes_dic = { nombre_alumno : importe } 
    for k,v in importes_dic.items() :
        hoja_ingresos[f'A{row}'] = k
        hoja_ingresos[f'B{row}'] = v
        padres = get_parents_by_student (k,ordenantes)
        hoja_ingresos[f'C{row}'] = padres
        # incrementa una fila en la hoja del libro de ingresos
        row += 1

    # Añade fila con los totales
    importes = list ( importes_dic.values() )
    total_ingresos = sum ( [ i[1] for i in importes ] )
    print(total_ingresos)
    hoja_ingresos[f'A{row}'] = 'Total'
    hoja_ingresos[f'B{row}'] = total_ingresos

    ### Format the cells of the workbook

    max_alumno_len = max ( \
        [len(alumno) for alumno in list(importes_dic.keys()) ] )
    max_padres_len = max ( \
        [len(padres) for padres in list(ordenantes.keys()) ] )

    print(max_alumno_len,max_padres_len)
    hoja_ingresos.column_dimensions['A'].width = max_alumno_len
    hoja_ingresos.column_dimensions['B'].width = len('Importe (€)')
    hoja_ingresos.column_dimensions['C'].width = max_padres_len


    # save the book
    nombre_libro_ingresos = 'ingresos_enero_21'
    libro_ingresos.save(f'{nombre_libro_ingresos}.xlsx')
    print(f'Libro guardado con el nombre: {nombre_libro_ingresos}.xlsx')


